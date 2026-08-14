#!/usr/bin/env python3
"""Build the data-collection / data-transfer workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()

H  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
B  = Font(bold=True, size=10)
N  = Font(size=10)
SM = Font(size=9, color='555555')
FH = PatternFill('solid', fgColor='1F3864')     # header navy
F1 = PatternFill('solid', fgColor='DEEBF7')     # meta
F2 = PatternFill('solid', fgColor='E2EFDA')     # focal items
F3 = PatternFill('solid', fgColor='FFF2CC')     # controls/marker
F4 = PatternFill('solid', fgColor='FCE4D6')     # objective
FW = PatternFill('solid', fgColor='FFC7CE')     # warning
FG = PatternFill('solid', fgColor='C6EFCE')     # good
thin = Side(style='thin', color='BFBFBF')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')


def head(ws, row, values, widths=None, fill=FH):
    for j, v in enumerate(values, 1):
        c = ws.cell(row, j, v); c.font = H; c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        c.border = BD
    ws.row_dimensions[row].height = 30
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[CL(j)].width = w


def rows(ws, start, data, wrapcols=()):
    for i, r in enumerate(data):
        for j, v in enumerate(r, 1):
            c = ws.cell(start + i, j, v); c.font = N; c.border = BD
            c.alignment = WRAP if j in wrapcols else Alignment(vertical='top')
    return start + len(data)


# =====================================================================
# 00 README
# =====================================================================
ws = wb.active; ws.title = '00_README'
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 118
ws['B1'] = 'Data workbook — circular accounting capability study'
ws['B1'].font = Font(bold=True, size=14)
readme = [
 ('', ''),
 ('SECTION', 'HOW TO USE THIS WORKBOOK'),
 ('', ''),
 ('1', 'Read 01_SampleSize first. It tells you what your achieved n allows you to claim, and it '
       'flags a power problem in the manuscript\'s current Approach A that you should fix before submission.'),
 ('2', 'Enter data on sheet 04_DATA only. One row per respondent. Do not insert, delete or reorder '
       'columns — the R script in analysis_seminr.R matches these names exactly.'),
 ('3', 'Column meanings are in 02_Codebook. Item wordings are in 03_Items. Categorical codes are in 05_ValueLabels.'),
 ('4', 'Track fieldwork progress in 06_SampleTracker. The response accounting it produces is required '
       'for Table 9 of the manuscript — reviewers will ask for it.'),
 ('5', '07_QualityChecks runs live per-row diagnostics as you enter data. Fix problems during fieldwork, '
       'not afterwards.'),
 ('6', 'When complete, save 04_DATA as data.csv (UTF-8) in the same folder as analysis_seminr.R and run it.'),
 ('', ''),
 ('SECTION', 'CRITICAL — CHECK THESE BEFORE YOU ANALYSE ANYTHING'),
 ('', ''),
 ('a', 'CAC_GLOBAL and CVCP_GLOBAL are single-item global measures of the two formative higher-order '
       'constructs. Redundancy analysis is impossible without them and CANNOT be done after the fact. '
       'If you did not field them, leave the columns empty and report the failure to establish convergent '
       'validity of the formative constructs as a limitation. Do not omit it silently.'),
 ('b', 'Check CGQ dispersion the moment you have the data (07_QualityChecks computes it). The entire design '
       'rests on governance quality varying widely within one network and one regulatory environment. '
       'If the SD is small, the interaction cannot be detected, and you must say so BEFORE reporting any null.'),
 ('c', 'All CVCP items must be worded so that HIGHER = BETTER. The loss items should read "reduction in…". '
       'If any was fielded as "quantity of waste", reverse-code it and record that you did.'),
 ('d', 'IS2 and IS3 are the double-barrelled items your Section 5.5 flags as below threshold and never '
       'revised. If they were fielded as written, report their psychometrics rather than quietly dropping them.'),
 ('e', 'MK1–MK3 must be Miller and Simmering\'s (2023) attitude-toward-blue items reproduced verbatim. '
       'A paraphrased marker forfeits the warrant for using a validated one.'),
 ('', ''),
 ('SECTION', 'WHAT NOT TO DO'),
 ('', ''),
 ('x', 'Do not delete cases flagged as long-string, fast-completion or multivariate outliers. These are '
       'flag-and-test, not delete. Deleting them without reporting the sensitivity analysis is the most '
       'common source of an undisclosed specification search in this literature.'),
 ('x', 'Do not add a CAC → CVCP path "to see if it is significant". H9\'s claim is that the effect is '
       'entirely indirect. If you estimate it you must report it, and the theoretical claim changes.'),
 ('x', 'Do not run Harman\'s single-factor test. Your Section 6 pre-specifies that it is not used.'),
 ('x', 'Do not run the producer-versus-processor multi-group comparison. With ~70 processors it needs '
       'β ≥ 0.30 to be detectable, and your manuscript already states it will not be conducted and why.'),
 ('x', 'Do not judge the interaction effect size against Cohen\'s 0.02/0.15/0.35. Use the interaction-specific '
       'benchmarks 0.005/0.01/0.025, as Table 8 pre-specifies.'),
]
r = 3
for tag, txt in readme:
    a = ws.cell(r, 1, tag); b = ws.cell(r, 2, txt)
    if tag == 'SECTION':
        a.value = ''; b.font = Font(bold=True, size=11, color='1F3864')
    elif tag == 'x':
        a.value = '✗'; a.font = Font(bold=True, color='C00000'); b.font = N
    elif tag in ('a', 'b', 'c', 'd', 'e'):
        a.font = Font(bold=True, color='BF8F00'); b.font = N
    else:
        a.font = B; b.font = N
    b.alignment = WRAP
    ws.row_dimensions[r].height = max(15, 13 * (1 + len(txt) // 105))
    r += 1

# =====================================================================
# 01 SAMPLE SIZE
# =====================================================================
ws = wb.create_sheet('01_SampleSize')
ws['A1'] = 'Required number of respondents'; ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = ('Figures recomputed from the non-central F distribution; they reproduce Table 6 of the '
            'manuscript exactly (Approach A f²=0.05 at power .90 → 434; minimum 346≈350 at power .80; '
            'Approach B β=0.11 → 511).')
ws['A2'].font = SM; ws['A2'].alignment = WRAP
ws.merge_cells('A2:F2'); ws.row_dimensions[2].height = 30

ws['A4'] = 'A. The design target (unchanged from the manuscript)'; ws['A4'].font = Font(bold=True, size=11, color='1F3864')
head(ws, 5, ['Approach', 'Basis', 'Requirement', 'Status', '', ''],
     [30, 62, 14, 26, 14, 14])
rows(ws, 6, [
 ['A. A-priori power', 'F-test for R² deviation from zero; 11 predictors on CVCI; f²=0.05; power .90; α=.05', 434, 'satisfied at n≥434', '', ''],
 ['B. Inverse square root', 'Kock and Hadaya (2018); smallest expected path β=0.11; power .80', 511, 'BINDING', '', ''],
 ['C. Model complexity', '≈96 free parameters at 5:1 for the covariance-based robustness check', 480, 'satisfied at n≥480', '', ''],
 ['D. Multi-group analysis', 'Two groups each estimable at β≈0.15–0.20', '490–550', 'satisfied at n≥490', '', ''],
 ['ADOPTED TARGET', '', 520, '', '', ''],
 ['Minimum acceptable', 'Approach A at power .80 (computed: 346)', 350, '', '', ''],
 ['Recommended', '', 450, '', '', ''],
], wrapcols=(2,))
for rr in (10,):
    for j in range(1, 5): ws.cell(rr, j).font = B

ws['A15'] = 'B. What your achieved n actually buys you'; ws['A15'].font = Font(bold=True, size=11, color='1F3864')
ws['A16'] = 'Find your row. This is the honest basis for what you can claim.'; ws['A16'].font = SM
head(ws, 17, ['Achieved n', 'Smallest detectable path β (power .80)',
              'Smallest detectable interaction f² (power .80)',
              'Approach A power at f²=0.05', 'MGA viable?', 'Verdict'],
     [12, 30, 32, 22, 16, 34])
sz = [
 (200, 0.176, 0.0398, 0.497, 'no',  'Insufficient. Below the stated minimum.'),
 (250, 0.157, 0.0318, 0.620, 'no',  'Insufficient.'),
 (300, 0.144, 0.0265, 0.723, 'no',  'Below minimum. Interaction test weak.'),
 (350, 0.133, 0.0226, 0.806, 'yes', 'Minimum acceptable. State the limitation.'),
 (400, 0.124, 0.0198, 0.868, 'yes', 'Acceptable.'),
 (450, 0.117, 0.0175, 0.913, 'yes', 'Recommended level reached.'),
 (480, 0.113, 0.0165, 0.933, 'yes', 'Approach C satisfied.'),
 (511, 0.110, 0.0155, 0.949, 'yes', 'Binding Approach B satisfied.'),
 (520, 0.109, 0.0152, 0.953, 'yes', 'DESIGN TARGET MET.'),
 (600, 0.101, 0.0132, 0.978, 'yes', 'Comfortable.'),
 (700, 0.094, 0.0112, 0.992, 'yes', 'Comfortable; medium interaction detectable.'),
]
rr = 18
for n_, b_, f_, p_, m_, v_ in sz:
    for j, v in enumerate([n_, b_, f_, p_, m_, v_], 1):
        c = ws.cell(rr, j, v); c.font = N; c.border = BD
        c.alignment = WRAP if j == 6 else Alignment(horizontal='center', vertical='top')
    ws.cell(rr, 2).number_format = '0.000'
    ws.cell(rr, 3).number_format = '0.0000'
    ws.cell(rr, 4).number_format = '0.000'
    fill = FG if n_ >= 511 else (FW if n_ < 350 else F3)
    for j in range(1, 7): ws.cell(rr, j).fill = fill
    if n_ == 520:
        for j in range(1, 7): ws.cell(rr, j).font = B
    rr += 1

ws['A31'] = 'C. A power problem in the manuscript you should fix'; ws['A31'].font = Font(bold=True, size=11, color='C00000')
ws['A32'] = (
 'Your Section 5.4 argues, correctly, that "the binding quantity is the interaction, not the largest path" and '
 'that "planning against a medium main effect would under-power the central test". Approach A then plans against '
 'f²=0.05 on the overall R² of CVCI — which is a MAIN-EFFECT quantity. Judged against the interaction-specific '
 'benchmarks your own Table 8 pre-specifies (0.005 small / 0.01 medium / 0.025 large), the design target of 520 '
 'gives power .95 for a LARGE interaction but only .62 for a MEDIUM one and .36 for a SMALL one. The argument is '
 'right; the number does not follow it through. A reviewer who checks will notice.')
ws['A32'].font = N; ws['A32'].alignment = WRAP
ws.merge_cells('A32:F32'); ws.row_dimensions[32].height = 62

head(ws, 34, ['Achieved n', 'Power at f²=0.005 (small)', 'Power at f²=0.010 (medium)',
              'Power at f²=0.015', 'Power at f²=0.025 (large)', 'Smallest f² at power .80'],
     [12, 30, 32, 22, 16, 34])
pwr = [(350,.261,.463,.627,.839,.0226),(400,.292,.514,.686,.884,.0198),
       (450,.322,.562,.736,.917,.0175),(480,.340,.590,.764,.933,.0165),
       (511,.358,.617,.789,.946,.0155),(520,.363,.624,.796,.949,.0152),
       (600,.409,.686,.850,.972,.0132)]
rr = 35
for row_ in pwr:
    for j, v in enumerate(row_, 1):
        c = ws.cell(rr, j, v); c.font = B if row_[0] == 520 else N; c.border = BD
        c.alignment = Alignment(horizontal='center')
        if j > 1: c.number_format = '0.000' if j < 6 else '0.0000'
        if j in (2, 3) and isinstance(v, float) and v < 0.8: c.fill = FW
        if j in (4, 5) and isinstance(v, float) and v >= 0.8: c.fill = FG
    rr += 1

ws['A43'] = 'The fix — add this sentence to Section 5.4, and honour it in Section 7'
ws['A43'].font = Font(bold=True, size=10, color='1F3864')
ws['A44'] = (
 '"Because the interaction is the binding quantity, we additionally report the minimum detectable interaction '
 'effect size. At the achieved sample of n = [___], the design attains power .80 for an interaction f² of [___] '
 'or larger, judged against the interaction-specific benchmarks of 0.005, 0.010 and 0.025. A true interaction '
 'below that value would not reliably be detected, and a null result is therefore interpreted accordingly."\n\n'
 'This costs you two sentences and converts a vulnerability into evidence of methodological care. If the '
 'interaction turns out null, it is also what prevents the null being over-read as evidence of absence.')
ws['A44'].font = N; ws['A44'].alignment = WRAP
ws.merge_cells('A44:F44'); ws.row_dimensions[44].height = 90

# =====================================================================
# Column definitions
# =====================================================================
SCALE = ('1–7', '7-point; 1 = not at all … 7 = to a very great extent')
CHG   = ('1–7', '7-point change; 1 = decreased greatly, 4 = no change, 7 = increased greatly')

COLS = []
COLS += [
 ('ID','Respondent identifier','text','unique','Required. Any unique code.',F1),
 ('DATE','Date of interview','date','yyyy-mm-dd','Required.',F1),
 ('ENUM','Enumerator identifier','text','','Required — needed for the enumerator-effects robustness check.',F1),
 ('DUR_min','Interview duration in minutes','integer','>0','Required — needed for the fast-completion flag.',F1),
 ('STRATUM','Sampling stratum','integer','1–6','See 05_ValueLabels. Required for Table 9 response accounting.',F1),
 ('TIER','Network position','integer','1–4','See 05_ValueLabels. Used as the MGA grouping variable, NOT as a covariate.',F1),
 ('SCR1','Keeps or supervises records?','integer','0/1','Screening. Must be 1 to be eligible.',F1),
 ('SCR2','Participates in pricing / disposal decisions?','integer','0/1','Screening. Must be 1 to be eligible.',F1),
 ('SCR3','Seasons in this role','integer','>=0','Screening. Must be >=3 (the outcome items are change over three seasons).',F1),
 ('SCR4','Self-described role','integer','1–6','Assigns STRATUM. See 05_ValueLabels.',F1),
]
def block(prefix, n, label, fill, note, scale=SCALE):
    return [(f'{prefix}{i}', f'{label} item {i}', 'integer', scale[0], note + ' ' + scale[1], fill)
            for i in range(1, n + 1)]
COLS += block('CAC', 9, 'Circular accounting capability','',
              'Focal. Items 1–3 measurement, 4–6 loss costing, 7–9 recovery valuation.')
COLS += block('CIT', 5, 'Circular information transparency','', 'Focal.')
COLS += block('CGQ', 6, 'Circular governance quality','', 'Focal — the moderator.')
COLS += block('CVCI', 6, 'Circular value-chain integration','', 'Focal — the focal endogenous construct.')
COLS += block('CVCP', 11,'Circular value creation performance','',
              'Focal. Items 1–4 economic, 5–8 environmental, 9–11 social. HIGHER = BETTER.', CHG)
COLS = [(a,b,c,d,e,F2 if f=='' else f) for a,b,c,d,e,f in COLS]
COLS += [(a,b,c,d,e,F3) for a,b,c,d,e,_ in block('DTC',3,'Digital traceability capability','',
              'Control and rival explanation. Critical to the smart-production test.')]
COLS += [(a,b,c,d,e,F3) for a,b,c,d,e,_ in block('IS',3,'Institutional support','',
              'Control. IS2 and IS3 are the double-barrelled items flagged in Section 5.5.')]
COLS += [(a,b,c,d,e,F3) for a,b,c,d,e,_ in block('MK',3,'Marker (attitude toward the colour blue)','',
              'Method-bias marker. Miller and Simmering (2023), verbatim. Excluded from the structural model.')]
COLS += [
 ('CAC_GLOBAL','Global single-item measure of circular accounting capability','integer','1–7',
  'REQUIRED FOR REDUNDANCY ANALYSIS of the formative HOC. Cannot be added later. Leave blank if not fielded and declare the limitation.',F3),
 ('CVCP_GLOBAL','Global single-item measure of circular value creation','integer','1–7',
  'REQUIRED FOR REDUNDANCY ANALYSIS of the formative HOC. Cannot be added later.',F3),
 ('SIZE_n','Persons engaged, incl. unpaid family labour','integer','>=1','Control. Script derives SIZE_ln = log(1+n).',F3),
 ('AGE_yr','Years the enterprise has operated','integer','>=0','Control. Script derives AGE_ln.',F3),
 ('VOL_t','Tonnes handled last season','decimal','>=0','Control. Script derives VOL_ln.',F3),
 ('EDU','Key informant education','integer','1–6','Control. See 05_ValueLabels.',F3),
 ('AFF','Producer-organisation / cooperative member','integer','0/1','Control and MGA grouping variable.',F3),
 ('OBJ1','Loss ratio from records: rejected or lost ÷ taken in','decimal','0–1','Objective subsample only; blank otherwise.',F4),
 ('OBJ2_qty','By-product quantity sold last season (tonnes)','decimal','>=0','Objective subsample only.',F4),
 ('OBJ2_val','By-product value realised last season (currency)','decimal','>=0','Objective subsample only.',F4),
 ('OBJ3','Realised price ÷ district modal price','decimal','>0','Objective subsample only.',F4),
 ('NOTES','Enumerator notes','text','','Optional. Not analysed.',F1),
]

# =====================================================================
# 02 CODEBOOK
# =====================================================================
ws = wb.create_sheet('02_Codebook')
ws['A1'] = 'Codebook — 68 columns, in the exact order required by analysis_seminr.R'
ws['A1'].font = Font(bold=True, size=13)
head(ws, 3, ['#','Column name','Label','Type','Allowed values','Notes'],
     [5, 16, 46, 10, 14, 74])
rr = 4
for i, (name, label, typ, allowed, note, fill) in enumerate(COLS, 1):
    for j, v in enumerate([i, name, label, typ, allowed, note], 1):
        c = ws.cell(rr, j, v); c.font = N; c.border = BD; c.fill = fill
        c.alignment = WRAP if j in (3, 6) else Alignment(vertical='top')
    if 'REQUIRED FOR REDUNDANCY' in note:
        for j in range(1, 7): ws.cell(rr, j).fill = FW
    rr += 1
ws.freeze_panes = 'B4'
ws.auto_filter.ref = f'A3:F{rr-1}'

# =====================================================================
# 03 ITEMS
# =====================================================================
ws = wb.create_sheet('03_Items')
ws['A1'] = 'The 46-item instrument'; ws['A1'].font = Font(bold=True, size=13)
ws['A2'] = ('RECONSTRUCTION — the instrument table is not present in JCLP_11_08_26.docx. These wordings satisfy '
            'every constraint the manuscript states (counts, dimensions, the 22/11/10/3 source split, and the item '
            'labels named in Section 5.5), but they are NOT retrieved from your file. Replace column F with your '
            'fielded wordings and keep the IDs aligned.')
ws['A2'].font = Font(size=9, color='C00000'); ws['A2'].alignment = WRAP
ws.merge_cells('A2:G2'); ws.row_dimensions[2].height = 46
head(ws, 4, ['ID','Construct','Dimension','Source','R3 reworded?','Item wording','Watch for'],
     [10, 12, 34, 8, 12, 88, 40])

ITEMS = [
 ('CAC1','CAC','Material and resource-flow measurement','AS','','We keep a written or electronic record of the quantity of raw material we take in each season.',''),
 ('CAC2','CAC','Material and resource-flow measurement','AS','yes','We write down how much of what we handle is rejected, downgraded or lost.','The one item quoted in your manuscript.'),
 ('CAC3','CAC','Material and resource-flow measurement','AS','','We record what happens to the material that does not become saleable product — where it goes and who takes it.',''),
 ('CAC4','CAC','Loss and waste costing','AS','','We work out what the material we lose or reject cost us to buy or to produce.',''),
 ('CAC5','CAC','Loss and waste costing','AS','','We know what the rejected or lost material would have earned had it been sold as product.',''),
 ('CAC6','CAC','Loss and waste costing','ND','','We keep the cost of losses separate from our general operating costs rather than absorbing it into overall cost.',''),
 ('CAC7','CAC','Recovery and circular investment evaluation','ND','','Before deciding whether to recover a by-product or residue, we estimate what it could earn.','New dimension; expert disagreement concentrated here.'),
 ('CAC8','CAC','Recovery and circular investment evaluation','ND','','We compare the cost of recovering a residue against the cost of disposing of it.','New dimension.'),
 ('CAC9','CAC','Recovery and circular investment evaluation','ND','','When we consider equipment or capacity for processing residues, we work out whether the investment would pay back.','New dimension.'),
 ('CIT1','CIT','—','AV','','Partners we deal with can see how much material is rejected or lost at our stage.',''),
 ('CIT2','CIT','—','AV','','We can see how much material is rejected or lost at the stages before and after us.',''),
 ('CIT3','CIT','—','AV','yes','Information about recoverable residues and by-products reaches the people who could use them, in time for them to act.','Reached 0.975 after rewording.'),
 ('CIT4','CIT','—','AS','','Information about material flows and losses in this chain is accurate and reliable enough to base a decision on.',''),
 ('CIT5','CIT','—','ND','','We can see how the money paid for the final product is divided among the actors in this chain.','Price rather than material information — most likely to behave differently.'),
 ('CGQ1','CGQ','Responsibility allocation','ND','yes','It is clear which party is answerable when material is lost or spoiled between us and our partners.','The construct\'s diagnostic core.'),
 ('CGQ2','CGQ','Joint monitoring and verification','AV','','Our partners and we jointly check the quantities and quality recorded when material changes hands.',''),
 ('CGQ3','CGQ','Codified standards','AS','','What counts as an acceptable level of loss is agreed and recorded with our partners.',''),
 ('CGQ4','CGQ','Participatory decision-making','AV','','We are able to take part in decisions that affect how material and losses are handled in this chain.',''),
 ('CGQ5','CGQ','Participatory decision-making','AV','','Our views are taken into account when the terms of trade in this chain are set.',''),
 ('CGQ6','CGQ','Enforceable commitment','AV','yes','Commitments our partners make about handling, collection or payment are honoured, and there are consequences if they are not.','Low ρA here attenuates the interaction directly.'),
 ('CVCI1','CVCI','—','AV','','We plan jointly with partners how material that would otherwise be wasted will be collected and moved.',''),
 ('CVCI2','CVCI','—','AV','yes','We coordinate the timing of our operations with partners so that recoverable material is handled before it deteriorates.',''),
 ('CVCI3','CVCI','—','AV','','We share equipment, storage or transport with partners for handling residues and by-products.',''),
 ('CVCI4','CVCI','—','AV','','We and our partners exchange the information needed to route recoverable material to where it can be used.','Watch for HTMT overlap with CIT.'),
 ('CVCI5','CVCI','—','AS','','We agree with partners in advance how the proceeds from recovered material will be shared.',''),
 ('CVCI6','CVCI','—','AV','','We work together with partners to find new uses for material that used to be discarded.',''),
 ('CVCP1','CVCP','Economic','AV','','Income earned from selling by-products or residues that were previously discarded.',''),
 ('CVCP2','CVCP','Economic','AV','','Savings from reusing material internally instead of buying new.',''),
 ('CVCP3','CVCP','Economic','AS','','The price we realise for our main product.',''),
 ('CVCP4','CVCP','Economic','AS','','Reduction in the cost of disposing of waste.',''),
 ('CVCP5','CVCP','Environmental','AV','','Reduction in the quantity of material leaving our operation as waste.',''),
 ('CVCP6','CVCP','Environmental','AV','','Reduction in the share of what we handle that is lost or spoiled.',''),
 ('CVCP7','CVCP','Environmental','AS','','Reduction in the energy and water used per unit of product.',''),
 ('CVCP8','CVCP','Environmental','ND','','Reduction in the quantity of residues sent to landfill, burned or dumped.',''),
 ('CVCP9','CVCP','Social','ND','','The share of the value from recovered material that reaches the producers who supplied it.','Weak SOC weight is a FINDING, not a failure.'),
 ('CVCP10','CVCP','Social','ND','','Employment or paid work created by handling and processing recovered material.',''),
 ('CVCP11','CVCP','Social','ND','','Fairness with which the gains from reducing losses are shared among the actors in this chain.',''),
 ('DTC1','DTC (control)','—','AV','','We use digital tools — mobile apps, spreadsheets or software — to record the quantities we handle.','If β(CAC→CIT) collapses when DTC enters, the paper\'s premise fails.'),
 ('DTC2','DTC (control)','—','AV','','Material we handle can be traced back to its source using our records or codes.',''),
 ('DTC3','DTC (control)','—','AV','','We can exchange records of quantities and quality with our partners electronically.',''),
 ('IS1','IS (control)','—','AV','yes','Government or industry bodies provide us with useful guidance on reducing losses and using residues.',''),
 ('IS2','IS (control)','—','AV','no','Training and technical advice on handling and processing residues are available to us.','DOUBLE-BARRELLED, below threshold, never revised.'),
 ('IS3','IS (control)','—','AV','no','Financial support or subsidy is available for equipment that would let us process residues.','DOUBLE-BARRELLED, below threshold, never revised.'),
 ('MK1','MK (marker)','—','MK','no','[Reproduce verbatim from Miller and Simmering 2023 — do not paraphrase]','Low relevance is the QUALIFYING property of a marker.'),
 ('MK2','MK (marker)','—','MK','','[Reproduce verbatim from Miller and Simmering 2023]',''),
 ('MK3','MK (marker)','—','MK','yes','[Reproduce verbatim from Miller and Simmering 2023]',''),
]
rr = 5
for it in ITEMS:
    for j, v in enumerate(it, 1):
        c = ws.cell(rr, j, v); c.font = N; c.border = BD
        c.alignment = WRAP if j in (3, 6, 7) else Alignment(vertical='top', horizontal='center')
        c.fill = F2 if it[1] in ('CAC','CIT','CGQ','CVCI','CVCP') else F3
    if it[4] == 'no': ws.cell(rr, 5).fill = FW
    if it[4] == 'yes': ws.cell(rr, 5).fill = FG
    rr += 1
ws.freeze_panes = 'A5'
ws.cell(rr+1, 1, 'Source: AV = adapted from a validated scale (22). AS = validated item structure adapted to circular '
        'resource-flow content (11). ND = newly developed (10). MK = validated marker administered verbatim (3). '
        'Total 46, reproducing the split in Section 5.5.').font = SM

# =====================================================================
# 04 DATA
# =====================================================================
ws = wb.create_sheet('04_DATA')
for j, (name, label, typ, allowed, note, fill) in enumerate(COLS, 1):
    c = ws.cell(1, j, name); c.font = H; c.fill = FH
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = BD
    c.comment = Comment(f'{label}\nType: {typ}\nAllowed: {allowed}\n\n{note}', 'Codebook', height=150, width=320)
    ws.column_dimensions[CL(j)].width = max(9, min(14, len(name) + 3))
ws.row_dimensions[1].height = 34
ws.freeze_panes = 'A2'

NAMES = [c[0] for c in COLS]
def rng(nm, last=1001):
    j = NAMES.index(nm) + 1
    return f'{CL(j)}2:{CL(j)}{last}'

dv7 = DataValidation(type='whole', operator='between', formula1=1, formula2=7,
                     allow_blank=True, showErrorMessage=True,
                     errorTitle='Out of range', error='Seven-point scale: enter a whole number 1–7.')
ws.add_data_validation(dv7)
for nm in [c[0] for c in COLS if c[3] == '1–7']:
    dv7.add(rng(nm))

dv01 = DataValidation(type='whole', operator='between', formula1=0, formula2=1, allow_blank=True,
                      showErrorMessage=True, errorTitle='Binary', error='Enter 0 or 1.')
ws.add_data_validation(dv01)
for nm in ('SCR1','SCR2','AFF'): dv01.add(rng(nm))

for nm, lo, hi, ttl in (('STRATUM',1,6,'Stratum 1–6'), ('TIER',1,4,'Tier 1–4'),
                        ('SCR4',1,6,'Role 1–6'), ('EDU',1,6,'Education 1–6')):
    dv = DataValidation(type='whole', operator='between', formula1=lo, formula2=hi, allow_blank=True,
                        showErrorMessage=True, errorTitle=ttl, error=f'Enter a whole number {lo}–{hi}. See 05_ValueLabels.')
    ws.add_data_validation(dv); dv.add(rng(nm))

dvpos = DataValidation(type='decimal', operator='greaterThanOrEqual', formula1=0, allow_blank=True,
                       showErrorMessage=True, errorTitle='Must be non-negative', error='Enter a value >= 0.')
ws.add_data_validation(dvpos)
for nm in ('SCR3','DUR_min','SIZE_n','AGE_yr','VOL_t','OBJ2_qty','OBJ2_val','OBJ3'): dvpos.add(rng(nm))

dvratio = DataValidation(type='decimal', operator='between', formula1=0, formula2=1, allow_blank=True,
                         showErrorMessage=True, errorTitle='Ratio', error='Loss ratio must be between 0 and 1.')
ws.add_data_validation(dvratio); dvratio.add(rng('OBJ1'))

# =====================================================================
# 05 VALUE LABELS
# =====================================================================
ws = wb.create_sheet('05_ValueLabels')
ws['A1'] = 'Categorical coding'; ws['A1'].font = Font(bold=True, size=13)
head(ws, 3, ['Variable','Code','Label','Note'], [14, 8, 52, 66])
VL = []
for code, lab, note in [
 (1,'Producers, non-affiliated','Target 180 (34.6%). MGA group: upstream.'),
 (2,'Producers, producer-organisation members','Target 120 (23.1%). MGA group: upstream with collective governance.'),
 (3,'Aggregators, traders, commission agents','Target 80 (15.4%). The information-asymmetry tier.'),
 (4,'Processors and common facility centres','Target 70 (13.5%). Manufacturing and recovery actors; deliberately over-sampled.'),
 (5,'Retailers and institutional buyers','Target 40 (7.7%). Downstream price-formation tier.'),
 (6,'Network support institutions','Target 30 (5.8%). EXCLUDED from group analysis.')]:
    VL.append(('STRATUM', code, lab, note))
for code, lab, note in [
 (1,'Upstream producer','Strata 1–2'),(2,'Midstream intermediary','Stratum 3'),
 (3,'Processing / recovery','Stratum 4'),(4,'Downstream buyer','Stratum 5')]:
    VL.append(('TIER', code, lab, note))
VL.append(('TIER','(blank)','Support institutions','Stratum 6 — leave TIER blank; excluded from MGA.'))
for code, lab in [(1,'No formal schooling'),(2,'Primary'),(3,'Secondary'),
                  (4,'Higher secondary'),(5,'Graduate'),(6,'Postgraduate')]:
    VL.append(('EDU', code, lab, ''))
VL += [('AFF',0,'Not a member of a producer organisation or cooperative',''),
       ('AFF',1,'Member of a producer organisation or cooperative','MGA grouping variable.'),
       ('SCR1',0,'No — not eligible','Exclude the case.'),
       ('SCR1',1,'Yes — keeps or supervises records',''),
       ('SCR2',0,'No — not eligible','Exclude the case.'),
       ('SCR2',1,'Yes — participates in pricing / disposal decisions',''),
       ('SCR4','1–6','Same coding as STRATUM','Self-described role; used to assign STRATUM.')]
rows(ws, 4, VL, wrapcols=(3, 4))

# =====================================================================
# 06 SAMPLE TRACKER
# =====================================================================
ws = wb.create_sheet('06_SampleTracker')
ws['A1'] = 'Fieldwork tracker and response accounting (feeds Table 9)'
ws['A1'].font = Font(bold=True, size=13)
ws['A2'] = 'Enter columns C–G. Everything else calculates. Reviewers will ask for this reconciliation.'
ws['A2'].font = SM
head(ws, 4, ['Stratum','Target','Contacted','Refused','Screened out','Incomplete','Valid',
             'Response rate','% of achieved','Shortfall vs target'],
     [40, 9, 11, 9, 13, 12, 9, 12, 13, 18])
STR = [('1 Producers, non-affiliated',180),('2 Producers, PO members',120),
       ('3 Aggregators, traders, agents',80),('4 Processors and CFCs',70),
       ('5 Retailers and institutional buyers',40),('6 Network support institutions',30)]
rr = 5
for lab, tgt in STR:
    ws.cell(rr,1,lab).font=N; ws.cell(rr,2,tgt).font=N
    for j in range(3,8): ws.cell(rr,j).fill=F1
    ws.cell(rr,8,f'=IF(C{rr}=0,"",G{rr}/C{rr})').number_format='0.0%'
    ws.cell(rr,9,f'=IF($G$11=0,"",G{rr}/$G$11)').number_format='0.0%'
    ws.cell(rr,10,f'=G{rr}-B{rr}')
    for j in range(1,11): ws.cell(rr,j).border=BD; ws.cell(rr,j).font=N
    rr+=1
ws.cell(11,1,'TOTAL').font=B
for j,col in zip(range(2,8),'BCDEFG'):
    ws.cell(11,j,f'=SUM({col}5:{col}10)').font=B
ws.cell(11,8,'=IF(C11=0,"",G11/C11)').number_format='0.0%'
ws.cell(11,10,'=G11-B11').font=B
for j in range(1,11): ws.cell(11,j).border=BD; ws.cell(11,j).fill=F3

ws['A13']='Achieved n verdict'; ws['A13'].font=Font(bold=True,size=11,color='1F3864')
verdict=[('Achieved valid n','=G11'),
 ('Smallest detectable path β at power .80','=IF(G11=0,"",2.486/SQRT(G11))'),
 ('Approach A satisfied (>=434)','=IF(G11=0,"",IF(G11>=434,"YES","NO"))'),
 ('Approach B satisfied — BINDING (>=511)','=IF(G11=0,"",IF(G11>=511,"YES","NO"))'),
 ('Approach C satisfied (>=480)','=IF(G11=0,"",IF(G11>=480,"YES","NO"))'),
 ('Above stated minimum (>=350)','=IF(G11=0,"",IF(G11>=350,"YES","BELOW MINIMUM"))'),
 ('MGA viable (smallest group >=100)','Check group sizes on 07_QualityChecks')]
rr=14
for lab,f in verdict:
    ws.cell(rr,1,lab).font=N; c=ws.cell(rr,2,f); c.font=B
    if 'β' in lab: c.number_format='0.000'
    for j in (1,2): ws.cell(rr,j).border=BD
    rr+=1
ws['A22']=('If Approach B is not satisfied, do not delete the derivation from Section 5.4. Report the achieved n '
           'against all four approaches and add the minimum-detectable-interaction sentence from 01_SampleSize. '
           'Computing power against the interaction rather than the largest path is a strength; concealing a '
           'shortfall is not.')
ws['A22'].font=SM; ws['A22'].alignment=WRAP; ws.merge_cells('A22:J22'); ws.row_dimensions[22].height=44

# =====================================================================
# 07 QUALITY CHECKS
# =====================================================================
ws = wb.create_sheet('07_QualityChecks')
ws['A1']='Live per-row diagnostics and construct dispersion'; ws['A1'].font=Font(bold=True,size=13)
ws['A2']=('Formulas reference 04_DATA rows 2:1001. Flags are FLAG-AND-TEST, not delete — except eligibility '
          'and full straight-lining, which are exclusions under the pre-specified rules.')
ws['A2'].font=SM; ws['A2'].alignment=WRAP; ws.merge_cells('A2:H2'); ws.row_dimensions[2].height=30

FOC_FIRST, FOC_LAST = CL(NAMES.index('CAC1')+1), CL(NAMES.index('CVCP11')+1)
head(ws,4,['Row','ID','Eligible?','Missing focal items','% missing','Straight-lining','Fast completion','Action'],
     [7,14,20,17,10,15,15,34])
for i in range(2,1002):
    r=i+3
    F=f"'04_DATA'!${FOC_FIRST}{i}:${FOC_LAST}{i}"
    ws.cell(r,1,i)
    ws.cell(r,2,f"=IF('04_DATA'!A{i}=\"\",\"\",'04_DATA'!A{i})")
    ws.cell(r,3,f"=IF('04_DATA'!A{i}=\"\",\"\",IF(AND('04_DATA'!{CL(NAMES.index('SCR1')+1)}{i}=1,"
                f"'04_DATA'!{CL(NAMES.index('SCR2')+1)}{i}=1,"
                f"'04_DATA'!{CL(NAMES.index('SCR3')+1)}{i}>=3),\"eligible\",\"EXCLUDE\"))")
    ws.cell(r,4,f'=IF(\'04_DATA\'!A{i}="","",37-COUNT({F}))')
    ws.cell(r,5,f'=IF(\'04_DATA\'!A{i}="","",(37-COUNT({F}))/37)').number_format='0%'
    ws.cell(r,6,f'=IF(\'04_DATA\'!A{i}="","",IF(AND(COUNT({F})=37,'
               f'COUNTIF({F},INDEX({F},1))=37),"EXCLUDE","ok"))')
    ws.cell(r,7,f"=IF('04_DATA'!A{i}=\"\",\"\",IF('04_DATA'!{CL(NAMES.index('DUR_min')+1)}{i}"
               f"<PERCENTILE('04_DATA'!${CL(NAMES.index('DUR_min')+1)}$2:${CL(NAMES.index('DUR_min')+1)}$1001,0.05),\"FLAG\",\"ok\"))")
    ws.cell(r,8,f'=IF(\'04_DATA\'!A{i}="","",IF(OR(C{r}="EXCLUDE",F{r}="EXCLUDE"),"Exclude under pre-specified rule",'
               f'IF(E{r}>0.15,"Exclude: >15% missing",IF(G{r}="FLAG","Retain; test in robustness battery","Retain"))))')
    for j in range(1,9): ws.cell(r,j).font=N; ws.cell(r,j).border=BD
ws.freeze_panes='A5'
for col in ('C','F','G','H'):
    ws.conditional_formatting  # placeholder; conditional fills below via simple rule
from openpyxl.formatting.rule import CellIsRule
for col in ('C','F'):
    ws.conditional_formatting.add(f'{col}5:{col}1004',
        CellIsRule(operator='equal', formula=['"EXCLUDE"'], fill=FW))
ws.conditional_formatting.add('G5:G1004', CellIsRule(operator='equal', formula=['"FLAG"'], fill=F3))

ws['J4']='Construct dispersion — CHECK CGQ FIRST'; ws['J4'].font=Font(bold=True,size=11,color='C00000')
ws.column_dimensions['J'].width=34
for cc,w_ in (('K',10),('L',10),('M',10),('N',10),('O',44)):
    ws.column_dimensions[cc].width=w_
for j2, lab2 in enumerate(['Construct','Mean','SD','Min','Max','Interpretation'], 10):
    c = ws.cell(5, j2, lab2); c.font = H; c.fill = FH; c.border = BD
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
for j,(nm,lab) in enumerate([('CAC','Circular accounting capability'),('CIT','Circular information transparency'),
        ('CGQ','Circular governance quality'),('CVCI','Circular value-chain integration'),
        ('CVCP','Circular value creation performance'),('DTC','Digital traceability capability')]):
    r=6+j
    first,last = {'CAC':('CAC1','CAC9'),'CIT':('CIT1','CIT5'),'CGQ':('CGQ1','CGQ6'),
                  'CVCI':('CVCI1','CVCI6'),'CVCP':('CVCP1','CVCP11'),'DTC':('DTC1','DTC3')}[nm]
    rg=f"'04_DATA'!${CL(NAMES.index(first)+1)}$2:${CL(NAMES.index(last)+1)}$1001"
    ws.cell(r,10,lab).font=B if nm=='CGQ' else N
    ws.cell(r,11,f'=IFERROR(AVERAGE({rg}),"")').number_format='0.000'
    ws.cell(r,12,f'=IFERROR(STDEV({rg}),"")').number_format='0.000'
    ws.cell(r,13,f'=IFERROR(MIN({rg}),"")')
    ws.cell(r,14,f'=IFERROR(MAX({rg}),"")')
    ws.cell(r,15,'THE DESIGN DEPENDS ON THIS SD BEING LARGE. If it is small or the range is narrow, the '
                 'interaction cannot be detected and you must say so before reporting any null.'
                 if nm=='CGQ' else '')
    ws.cell(r,15).alignment=WRAP; ws.cell(r,15).font=SM
    for j2 in range(10,16):
        ws.cell(r,j2).border=BD
        if nm=='CGQ': ws.cell(r,j2).fill=FW
ws.row_dimensions[8].height=44

ws['J14']='MGA group sizes (need >=100 per group)'; ws['J14'].font=Font(bold=True,size=11,color='1F3864')
AFFc=CL(NAMES.index('AFF')+1); TIERc=CL(NAMES.index('TIER')+1)
grp=[('AFF = 1 (PO member)',f"=COUNTIF('04_DATA'!${AFFc}$2:${AFFc}$1001,1)"),
     ('AFF = 0 (not affiliated)',f"=COUNTIF('04_DATA'!${AFFc}$2:${AFFc}$1001,0)"),
     ('TIER 1 upstream',f"=COUNTIF('04_DATA'!${TIERc}$2:${TIERc}$1001,1)"),
     ('TIER 2 midstream',f"=COUNTIF('04_DATA'!${TIERc}$2:${TIERc}$1001,2)"),
     ('TIER 3 processing / recovery',f"=COUNTIF('04_DATA'!${TIERc}$2:${TIERc}$1001,3)"),
     ('TIER 4 downstream',f"=COUNTIF('04_DATA'!${TIERc}$2:${TIERc}$1001,4)")]
for k,(lab,f) in enumerate(grp):
    r=15+k
    ws.cell(r,10,lab).font=N; ws.cell(r,11,f).font=B
    ws.cell(r,12,f'=IF(K{r}>=100,"viable","under-powered")').font=N
    for j2 in range(10,13): ws.cell(r,j2).border=BD
ws.cell(21,10,'Do NOT run the producer-versus-processor comparison: with ~70 processors it needs β ≥ 0.30. '
              'Your manuscript states it will not be conducted and why. Honour that.').font=SM
ws.cell(21,10).alignment=WRAP

wb.save('Circular_Accounting_DataTemplate.xlsx')
print('saved Circular_Accounting_DataTemplate.xlsx')
print('sheets:', wb.sheetnames)
print('DATA columns:', len(COLS))
print('items listed:', len(ITEMS))
